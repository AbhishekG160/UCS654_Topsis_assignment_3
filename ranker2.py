"""
=======================================================================
  TEXT CLASSIFICATION MODEL SELECTION USING TOPSIS
  Using real PyPI packages: TextBlob, VADER, Flair, Transformers, NLTK
=======================================================================

Setup (inside your virtual environment):
    pip install textblob vaderSentiment flair transformers torch nltk
        numpy pandas scikit-learn openpyxl
    python -m textblob.download_corpora
    python -m nltk.downloader vader_lexicon

Each model comes from a DIFFERENT PyPI package with its own pre-trained weights.
TOPSIS ranks them on Accuracy, F1, Precision, Recall, Speed, and Memory.
Results are exported to an Excel workbook with one sheet per weight combination.
"""

import time
import tracemalloc
import warnings
import importlib
import importlib.util
import numpy as np
import pandas as pd
from datetime import datetime
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# -----------------------------------------------------------------------
# SAMPLE DATASET  (label: 0=Negative, 1=Positive, 2=Neutral)
# -----------------------------------------------------------------------
SAMPLES = [
    ("I absolutely loved this movie, it was fantastic!", 1),
    ("The film was a complete waste of time, terrible acting.", 0),
    ("An average experience, nothing special about it.", 2),
    ("Outstanding performance by the lead actor, truly brilliant.", 1),
    ("Worst product I have ever bought, total disappointment.", 0),
    ("It was okay, not great but not bad either.", 2),
    ("Incredible storyline and beautiful cinematography!", 1),
    ("Boring, dull, and completely uninspiring.", 0),
    ("Decent enough for a one-time watch.", 2),
    ("The food was amazing, best meal I've had in years!", 1),
    ("Terrible service and the food was cold.", 0),
    ("The restaurant was fine, nothing memorable.", 2),
    ("Highly recommend this product, changed my life!", 1),
    ("Do not buy this, it broke after one day.", 0),
    ("It works as described, fairly straightforward.", 2),
    ("Such a heartwarming and uplifting story.", 1),
    ("Painful to watch, could not finish it.", 0),
    ("It was a standard, run-of-the-mill experience.", 2),
    ("Phenomenal soundtrack, the music was breathtaking.", 1),
    ("Mediocre at best, I expected much more from this.", 2),
    ("Disgraceful quality, I want a refund immediately.", 0),
    ("Surprisingly good for the price, very pleased.", 1),
    ("Not what I expected, but acceptable overall.", 2),
    ("Pure joy from start to finish, a masterpiece!", 1),
    ("Absolutely horrible, I regret every penny spent.", 0)
]

TEXTS  = [s[0] for s in SAMPLES]
LABELS = [s[1] for s in SAMPLES]


# -----------------------------------------------------------------------
# WEIGHT COMBINATIONS  — add more dicts here to test new scenarios
# -----------------------------------------------------------------------
WEIGHT_CONFIGS = {
    "Balanced": {
        "weights": {
            "Accuracy":            0.30,
            "F1-Score (weighted)": 0.25,
            "Precision":           0.15,
            "Recall":              0.15,
            "Inference Time (s)":  0.08,
            "Peak Memory (MB)":    0.07,
        },
        "description": "Balanced: Accuracy 30% | F1 25% | Precision 15% | Recall 15% | Time 8% | Memory 7%",
    },
    "Accuracy-Priority": {
        "weights": {
            "Accuracy":            0.45,
            "F1-Score (weighted)": 0.35,
            "Precision":           0.07,
            "Recall":              0.07,
            "Inference Time (s)":  0.03,
            "Peak Memory (MB)":    0.03,
        },
        "description": "Accuracy-Priority: Accuracy 45% | F1 35% | Precision 7% | Recall 7% | Time 3% | Memory 3%",
    },
    "Speed-Priority": {
        "weights": {
            "Accuracy":            0.05,
            "F1-Score (weighted)": 0.05,
            "Precision":           0.05,
            "Recall":              0.05,
            "Inference Time (s)":  0.45,
            "Peak Memory (MB)":    0.35,
        },
        "description": "Speed-Priority: Time 45% | Memory 35% | Accuracy 5% | F1 5% | Precision 5% | Recall 5%",
    },
}

BENEFIT = {
    "Accuracy":            True,
    "F1-Score (weighted)": True,
    "Precision":           True,
    "Recall":              True,
    "Inference Time (s)":  False,
    "Peak Memory (MB)":    False,
}


# -----------------------------------------------------------------------
# HELPERS
# -----------------------------------------------------------------------
def normalize_sentiment(raw: str) -> int:
    r = str(raw).lower()
    if any(w in r for w in ["pos", "1", "good", "happy", "joy"]):
        return 1
    elif any(w in r for w in ["neg", "0", "bad", "sad", "hate"]):
        return 0
    return 2

def hex_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


# -----------------------------------------------------------------------
# MODELS  (one per PyPI package)
# -----------------------------------------------------------------------
def run_textblob(texts):
    from textblob import TextBlob
    preds = []
    for text in texts:
        p = TextBlob(text).sentiment.polarity
        preds.append(1 if p > 0.1 else (0 if p < -0.1 else 2))
    return preds

def run_vader(texts):
    from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
    sia = SentimentIntensityAnalyzer()
    preds = []
    for text in texts:
        s = sia.polarity_scores(text)["compound"]
        preds.append(1 if s >= 0.05 else (0 if s <= -0.05 else 2))
    return preds

def run_flair(texts):
    from flair.models import TextClassifier
    from flair.data import Sentence
    clf = TextClassifier.load("en-sentiment")
    preds = []
    for text in texts:
        s = Sentence(text)
        clf.predict(s)
        label = s.labels[0].value if s.labels else "NEUTRAL"
        preds.append(normalize_sentiment(label))
    return preds

def run_transformers_zeroshot(texts):
    from transformers import pipeline
    clf = pipeline("zero-shot-classification", model="cross-encoder/nli-MiniLM2-L6-H768")
    preds = []
    for text in texts:
        r = clf(text, ["positive", "negative", "neutral"])
        preds.append(normalize_sentiment(r["labels"][0]))
    return preds

def run_transformers_sentiment(texts):
    from transformers import pipeline
    clf = pipeline("sentiment-analysis", model="distilbert-base-uncased-finetuned-sst-2-english")
    preds = []
    for text in texts:
        r = clf(text)[0]
        preds.append(2 if r["score"] < 0.65 else normalize_sentiment(r["label"]))
    return preds

def run_nltk(texts):
    import nltk
    from nltk.sentiment import SentimentIntensityAnalyzer
    nltk.download("vader_lexicon", quiet=True)
    sia = SentimentIntensityAnalyzer()
    preds = []
    for text in texts:
        s = sia.polarity_scores(text)["compound"]
        preds.append(1 if s >= 0.05 else (0 if s <= -0.05 else 2))
    return preds

MODELS = {
    "TextBlob":         (run_textblob,               ["textblob"]),
    "VADER":            (run_vader,                  ["vaderSentiment"]),
    "Flair":            (run_flair,                  ["flair"]),
    "Zero-Shot (HF)":   (run_transformers_zeroshot,  ["transformers", "torch"]),
    "DistilBERT (HF)":  (run_transformers_sentiment, ["transformers", "torch"]),
    "NLTK":             (run_nltk,                   ["nltk"]),
}


# -----------------------------------------------------------------------
# EVALUATION
# -----------------------------------------------------------------------
def evaluate_model(name, fn, required_pkgs, texts, labels):
    print(f"  [{name}]...", end=" ", flush=True)
    missing = [p for p in required_pkgs if not importlib.util.find_spec(p.split(".")[0])]
    if missing:
        print(f"SKIPPED — pip install {' '.join(missing)}")
        return None
    try:
        tracemalloc.start()
        t0 = time.time()
        preds = fn(texts)
        elapsed = time.time() - t0
        _, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        acc  = accuracy_score(labels, preds)
        f1   = f1_score(labels, preds, average="weighted", zero_division=0)
        prec = precision_score(labels, preds, average="weighted", zero_division=0)
        rec  = recall_score(labels, preds, average="weighted", zero_division=0)
        print(f"Acc={acc:.3f}  F1={f1:.3f}  Time={elapsed:.2f}s  Mem={peak/1e6:.1f}MB")
        return {
            "Accuracy":            round(acc, 4),
            "F1-Score (weighted)": round(f1, 4),
            "Precision":           round(prec, 4),
            "Recall":              round(rec, 4),
            "Inference Time (s)":  round(elapsed, 3),
            "Peak Memory (MB)":    round(peak / 1_048_576, 2),
        }
    except Exception as e:
        tracemalloc.stop()
        print(f"FAILED — {e}")
        return None


# -----------------------------------------------------------------------
# TOPSIS
# -----------------------------------------------------------------------
def topsis(df: pd.DataFrame, weights: dict, benefit: dict) -> pd.DataFrame:
    cols = list(df.columns)
    X    = df[cols].values.astype(float)
    norms = np.sqrt((X ** 2).sum(axis=0))
    norms[norms == 0] = 1e-9
    R = X / norms
    w = np.array([weights[c] for c in cols], dtype=float)
    w /= w.sum()
    V = R * w
    is_ben  = np.array([benefit[c] for c in cols])
    A_best  = np.where(is_ben, V.max(axis=0), V.min(axis=0))
    A_worst = np.where(is_ben, V.min(axis=0), V.max(axis=0))
    S_best  = np.sqrt(((V - A_best)  ** 2).sum(axis=1))
    S_worst = np.sqrt(((V - A_worst) ** 2).sum(axis=1))
    denom = S_best + S_worst
    denom[denom == 0] = 1e-9
    C = S_worst / denom
    out = df.copy()
    out["S+ (dist ideal)"] = S_best.round(5)
    out["S- (dist worst)"] = S_worst.round(5)
    out["TOPSIS Score"]    = C.round(5)
    out["TOPSIS Rank"]     = out["TOPSIS Score"].rank(ascending=False).astype(int)
    return out.sort_values("TOPSIS Rank")


# -----------------------------------------------------------------------
# EXCEL — individual scenario sheet
# -----------------------------------------------------------------------
def write_topsis_sheet(ws, topsis_df, raw_df, config_name, config_desc, weights):
    C_TITLE  = "1F3864"
    C_BLUE   = "2E75B6"
    C_HDR_M  = "D6E4F0"
    C_HDR_T  = "FDEBD0"
    C_GOLD   = "FFF9C4"
    C_SILVER = "ECEFF1"
    C_BRONZE = "EFEBE9"
    C_EVEN   = "F7FBFF"

    WHITE  = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    BOLD   = Font(bold=True, name="Arial", size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center")

    metrics  = list(raw_df.columns)
    top_cols = ["S+ (dist ideal)", "S- (dist worst)", "TOPSIS Score", "TOPSIS Rank"]
    all_cols = ["Model"] + metrics + top_cols
    n_cols   = len(all_cols)

    # Row 1 — title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, f"TOPSIS Ranking — {config_name}")
    c.font = Font(color="FFFFFF", bold=True, name="Arial", size=14)
    c.fill = hex_fill(C_TITLE); c.alignment = CENTER

    # Row 2 — description
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(2, 1, config_desc)
    c.font = Font(color="FFFFFF", italic=True, name="Arial", size=10)
    c.fill = hex_fill(C_BLUE); c.alignment = CENTER

    # Row 3 — weights row
    ws.cell(3, 1, "Criterion Weight").font = BOLD
    ws.cell(3, 1).alignment = LEFT
    for ci, col in enumerate(metrics, start=2):
        c = ws.cell(3, ci, f"{weights.get(col, 0)*100:.0f}%")
        c.font = Font(bold=True, name="Arial", size=10, color="1F3864")
        c.fill = hex_fill(C_HDR_M); c.alignment = CENTER
    for ci in range(2 + len(metrics), n_cols + 1):
        c = ws.cell(3, ci, "—")
        c.fill = hex_fill(C_HDR_T); c.alignment = CENTER

    # Row 4 — column headers
    for ci, col in enumerate(all_cols, start=1):
        c = ws.cell(4, ci, col)
        c.font = WHITE
        c.fill = hex_fill(C_BLUE if ci <= 1 + len(metrics) else C_TITLE)
        c.alignment = CENTER; c.border = thin_border()

    # Data rows
    medal_fills = {1: C_GOLD, 2: C_SILVER, 3: C_BRONZE}
    medal_emoji = {1: " 🥇", 2: " 🥈", 3: " 🥉"}

    for ri, (model_name, row) in enumerate(topsis_df.iterrows(), start=5):
        rank = int(row["TOPSIS Rank"])
        bg   = medal_fills.get(rank, C_EVEN if ri % 2 == 0 else "FFFFFF")

        c = ws.cell(ri, 1, model_name)
        c.font = Font(bold=(rank == 1), name="Arial", size=10)
        c.fill = hex_fill(bg); c.alignment = LEFT; c.border = thin_border()

        for ci, col in enumerate(metrics, start=2):
            val = row.get(col, "")
            c = ws.cell(ri, ci, val)
            c.fill = hex_fill(bg); c.alignment = CENTER
            c.border = thin_border()
            c.font = Font(name="Arial", size=10, bold=(rank == 1))
            if col in ("Accuracy", "F1-Score (weighted)", "Precision", "Recall"):
                c.number_format = "0.00%"
            elif col == "Inference Time (s)":
                c.number_format = '0.000"s"'
            elif col == "Peak Memory (MB)":
                c.number_format = '0.00"MB"'

        for ci, col in enumerate(top_cols, start=2 + len(metrics)):
            val = row.get(col, "")
            if col == "TOPSIS Rank":
                display = f"{int(val)}{medal_emoji.get(int(val), '')}"
                c = ws.cell(ri, ci, display)
            else:
                c = ws.cell(ri, ci, val)
                if col == "TOPSIS Score":
                    c.number_format = "0.0000"
                    c.font = Font(name="Arial", size=10, bold=(rank == 1),
                                  color="C0392B" if rank == 1 else "000000")
            c.fill = hex_fill(bg); c.alignment = CENTER; c.border = thin_border()
            if col != "TOPSIS Score":
                c.font = Font(name="Arial", size=10, bold=(rank == 1))

    # Column widths & row heights
    ws.column_dimensions["A"].width = 22
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 19
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 36


# -----------------------------------------------------------------------
# EXCEL — summary sheet
# -----------------------------------------------------------------------
def write_summary_sheet(ws, all_topsis: dict, raw_df: pd.DataFrame):
    C_TITLE = "1F3864"
    C_BLUE  = "2E75B6"
    C_GREEN = "2E7D32"
    C_EVEN  = "F7FBFF"
    WHITE   = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    BOLD    = Font(bold=True, name="Arial", size=10)
    CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT    = Alignment(horizontal="left",   vertical="center")

    config_names = list(all_topsis.keys())
    metric_cols  = list(raw_df.columns)
    n_cols = 1 + len(metric_cols) + len(config_names) + 1  # model+metrics+ranks+best

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, "TOPSIS Sensitivity Analysis — All Weight Scenarios")
    c.font = Font(color="FFFFFF", bold=True, name="Arial", size=14)
    c.fill = hex_fill(C_TITLE); c.alignment = CENTER

    # Section sub-headers row 2
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1+len(metric_cols))
    c = ws.cell(2, 2, "Performance Metrics (raw)")
    c.font = WHITE; c.fill = hex_fill("37474F"); c.alignment = CENTER

    ws.merge_cells(start_row=2, start_column=2+len(metric_cols),
                   end_row=2, end_column=1+len(metric_cols)+len(config_names))
    c = ws.cell(2, 2+len(metric_cols), "TOPSIS Rank per Scenario")
    c.font = WHITE; c.fill = hex_fill(C_BLUE); c.alignment = CENTER

    c = ws.cell(2, n_cols, "Best\nRank")
    c.font = WHITE; c.fill = hex_fill(C_GREEN); c.alignment = CENTER

    # Column headers row 3
    headers = ["Model"] + metric_cols + config_names + ["Best Rank"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(3, ci, h)
        c.font = WHITE; c.fill = hex_fill(C_BLUE)
        c.alignment = CENTER; c.border = thin_border()

    # Weight sub-row (row 4) — show weights for rank columns
    ws.cell(4, 1, "Criterion Weight →").font = BOLD
    ws.cell(4, 1).alignment = LEFT
    for ci, col in enumerate(metric_cols, start=2):
        ws.cell(4, ci, "").fill = hex_fill("E3F2FD")
    for ci, cfg_name in enumerate(config_names, start=2+len(metric_cols)):
        desc = WEIGHT_CONFIGS[cfg_name]["description"]
        c = ws.cell(4, ci, desc)
        c.font = Font(name="Arial", size=8, italic=True, color="37474F")
        c.fill = hex_fill("E8EAF6"); c.alignment = CENTER
        c.border = thin_border()
    ws.row_dimensions[4].height = 40

    # Data rows (starting row 5)
    for ri, model in enumerate(raw_df.index, start=5):
        bg = C_EVEN if ri % 2 == 0 else "FFFFFF"
        c = ws.cell(ri, 1, model)
        c.font = BOLD; c.fill = hex_fill(bg)
        c.alignment = LEFT; c.border = thin_border()

        # Raw metrics
        for ci, col in enumerate(metric_cols, start=2):
            val = raw_df.loc[model, col]
            c = ws.cell(ri, ci, val)
            c.fill = hex_fill(bg); c.alignment = CENTER
            c.border = thin_border()
            c.font = Font(name="Arial", size=10)
            if col in ("Accuracy", "F1-Score (weighted)", "Precision", "Recall"):
                c.number_format = "0.00%"
            elif col == "Inference Time (s)":
                c.number_format = '0.000"s"'
            elif col == "Peak Memory (MB)":
                c.number_format = '0.00"MB"'

        # Rank per scenario
        ranks = []
        for ci, cfg_name in enumerate(config_names, start=2+len(metric_cols)):
            rank_val = int(all_topsis[cfg_name].loc[model, "TOPSIS Rank"])
            topsis_score = float(all_topsis[cfg_name].loc[model, "TOPSIS Score"])
            ranks.append(rank_val)
            medal = {1: " 🥇", 2: " 🥈", 3: " 🥉"}.get(rank_val, "")
            c = ws.cell(ri, ci, f"{rank_val}{medal}  ({topsis_score:.4f})")
            c.alignment = CENTER; c.border = thin_border()
            c.font = Font(name="Arial", size=10, bold=(rank_val == 1))
            c.fill = hex_fill("FFF9C4" if rank_val == 1 else bg)

        # Best rank
        best = min(ranks)
        c = ws.cell(ri, n_cols, best)
        c.alignment = CENTER; c.border = thin_border()
        c.font = Font(name="Arial", size=11, bold=True,
                      color=C_GREEN if best == 1 else "000000")
        c.fill = hex_fill("C8E6C9" if best == 1 else bg)

    ws.column_dimensions["A"].width = 22
    for ci in range(2, 2 + len(metric_cols)):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    for ci in range(2 + len(metric_cols), n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 28
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 36


# -----------------------------------------------------------------------
# EXPORT
# -----------------------------------------------------------------------
def export_excel(raw_df, all_topsis, output_path):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary_sheet(ws_summary, all_topsis, raw_df)

    for config_name, topsis_df in all_topsis.items():
        cfg = WEIGHT_CONFIGS[config_name]
        ws = wb.create_sheet(title=config_name[:31])
        write_topsis_sheet(ws, topsis_df, raw_df, config_name,
                           cfg["description"], cfg["weights"])

    wb.save(output_path)
    print(f"\n  Excel report saved -> {output_path}")


# -----------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------
def main():
    print("=" * 70)
    print("  PYPI TEXT CLASSIFICATION BENCHMARK  ->  TOPSIS RANKING")
    print("=" * 70)
    print(f"  Dataset : {len(TEXTS)} sentences | 3 classes: neg / pos / neutral")
    print(f"  Models  : {len(MODELS)}\n")

    results = {}
    for name, (fn, pkgs) in MODELS.items():
        m = evaluate_model(name, fn, pkgs, TEXTS, LABELS)
        if m:
            results[name] = m

    if not results:
        print("\n[!] No models ran. Install required packages and retry.")
        return

    raw_df = pd.DataFrame(results).T
    print("\n" + "=" * 70)
    print("  RAW EVALUATION RESULTS")
    print("=" * 70)
    print(raw_df.round(4).to_string())

    all_topsis = {
        cfg_name: topsis(raw_df, cfg["weights"], BENEFIT)
        for cfg_name, cfg in WEIGHT_CONFIGS.items()
    }

    print("\n" + "=" * 70)
    print("  TOPSIS WINNERS PER SCENARIO")
    print("=" * 70)
    for cfg_name, t_df in all_topsis.items():
        winner = t_df.index[0]
        score  = t_df.loc[winner, "TOPSIS Score"]
        print(f"  {cfg_name:22s} ->  {winner}  (score={score:.4f})")

    print("\n" + "=" * 70)
    print("  SENSITIVITY TABLE  (ranks)")
    print("=" * 70)
    summary = pd.DataFrame({cfg: t["TOPSIS Rank"] for cfg, t in all_topsis.items()})
    print(summary.sort_values("Balanced").to_string())

    out_path = f"topsis_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_excel(raw_df, all_topsis, out_path)
    print("\nDone!")
    return out_path


if __name__ == "__main__":
    main()